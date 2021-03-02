#import Adrese
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
#from bs4 import BeautifulSoup
from time import sleep
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
#from helium import *
import pandas as pd
#import xlrd #pip install xlrd==1.2.0
from openpyxl import load_workbook

#options=Options()
#options.add_argument("--headless")
#driver=webdriver.Chrome(ChromeDriverManager().install(),options=options)
driver=webdriver.Chrome(ChromeDriverManager().install())

ls=[]
workbook = load_workbook(filename="Gradovi.xlsx")
sheet = workbook.active
for row in sheet.iter_rows(values_only=True):
            ls.append(row)

print("Obrada podataka započinje:")

list=[]
for item in ls:

    driver.get("https://www.a1.hr/INTERSHOP/web/WFS/A1-Shop-Site/hr_HR/a1-fixed/HRK/ViewRequisitionAvailability-View")
    sleep(2)

    #grad=driver.find_element_by_id('TechnologyAvailabilityForm_mjesto')
    #sleep(1)
    grad=WebDriverWait(driver, 20).until(
      EC.element_to_be_clickable((By.ID, 'TechnologyAvailabilityForm_mjesto')))
    for character in item[0]:
        grad.send_keys(character)
        sleep(0.3) # pause for 0.3 seconds
    grad.send_keys(Keys.RETURN)
    sleep(0.5)
    ulica = driver.find_element_by_id('TechnologyAvailabilityForm_Street')
    ulica.send_keys(item[1])
    sleep(0.5)
    broj = driver.find_element_by_id('TechnologyAvailabilityForm_StreetNumber')
    broj.send_keys(item[2])
    sleep(0.5)
    broj.send_keys(Keys.RETURN)
    sleep(3)
    #sign_in = driver.find_element_by_xpath('//*[@id="AvailabilityForm"]/fieldset[2]/div[3]/button').click()
    #sleep(3)

    data = {}
    if driver.current_url == "https://www.a1.hr/dostupnost/flatbox":
        flatbox = str(item)
        try:
            data["Flatbox"] = flatbox
        except:
            pass
    else:
        pass
    if driver.current_url == "https://www.a1.hr/dostupnost/none":
        #A="Paketi = "+ str(item)
        paketi = str(item)
        try:
            data["A1_paketi"] = paketi
        except:
            pass
    else:
        pass
    try:
        if driver.find_element_by_xpath('//*[contains(.,"Kabelske")]'):
            sleep(1)
            kablovska=str(item)
            try:
                data['Kablovska'] = kablovska
            except:
                pass
        else:
            pass
    except:
        pass
    try:
        if driver.find_element_by_xpath('//*[contains(.,"Optičke")]'):
            sleep(1)
            optika = str(item)
            try:
                data['Optika'] = optika
            except:
                pass
        else:
            pass
    except:
        pass

    list.append(data)
    print(data)
    #driver.back()

df = pd.DataFrame(list)
df.to_excel("A1_usluge.xlsx", index=False)
#df.to_csv("A1_usluge.csv", index=False)
print("Obrada podataka završena.")
driver.quit()
